import asyncio
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional, BinaryIO
from pathlib import Path
import tempfile
import uuid
import structlog
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

from jinja2 import Environment, BaseLoader
from weasyprint import HTML

from app.models.aws_account import AWSAccount
from app.services.cost_sync_service import cost_sync_service
from app.services.waste_detection_service import waste_detection_service
from app.services.recommendations_engine import recommendations_engine
from app.services.cache_service import cache_service
from app.db.base import get_database
from sqlalchemy.ext.asyncio import AsyncSession

logger = structlog.get_logger(__name__)


class TemplateLoader(BaseLoader):
    """Custom template loader for Jinja2"""

    def __init__(self, templates: Dict[str, str]):
        self.templates = templates

    def get_source(self, environment, template):
        if template in self.templates:
            source = self.templates[template]
            return source, None, lambda: True
        raise FileNotFoundError(f"Template {template} not found")


class ReportService:
    """Service for generating comprehensive cost optimization reports"""

    def __init__(self):
        self.temp_dir = Path(tempfile.gettempdir()) / "aws_cost_sentinel_reports"
        self.temp_dir.mkdir(exist_ok=True)
        self._setup_jinja_templates()

    def _setup_jinja_templates(self):
        """Set up Jinja2 templates for HTML reports"""
        templates = {
            'cost_report.html': '''
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>AWS Cost Report - {{ account_name }}</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 40px; }
        .header { text-align: center; border-bottom: 2px solid #333; padding-bottom: 20px; }
        .section { margin: 30px 0; }
        .metric { background: #f5f5f5; padding: 15px; margin: 10px 0; border-radius: 5px; }
        table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        th, td { padding: 10px; text-align: left; border: 1px solid #ddd; }
        th { background-color: #4CAF50; color: white; }
        .cost-value { font-weight: bold; color: #2E8B57; }
        .savings { color: #FF6347; font-weight: bold; }
        .chart-container { text-align: center; margin: 20px 0; }
    </style>
</head>
<body>
    <div class="header">
        <h1>AWS Cost Optimization Report</h1>
        <h2>{{ account_name }} ({{ account_id }})</h2>
        <p>Report Period: {{ start_date }} to {{ end_date }}</p>
        <p>Generated: {{ generated_at }}</p>
    </div>

    <div class="section">
        <h2>Executive Summary</h2>
        <div class="metric">
            <h3>Total Cost: <span class="cost-value">${{ "%.2f"|format(summary.total_cost) }}</span></h3>
        </div>
        <div class="metric">
            <h3>Potential Savings: <span class="savings">${{ "%.2f"|format(summary.potential_savings) }}</span></h3>
        </div>
        <div class="metric">
            <h3>Waste Items Found: {{ summary.waste_items_count }}</h3>
        </div>
        <div class="metric">
            <h3>Active Recommendations: {{ summary.recommendations_count }}</h3>
        </div>
    </div>

    <div class="section">
        <h2>Cost Breakdown by Service</h2>
        <table>
            <thead>
                <tr><th>Service</th><th>Cost ($)</th><th>Percentage</th></tr>
            </thead>
            <tbody>
                {% for service in cost_breakdown %}
                <tr>
                    <td>{{ service.service_name }}</td>
                    <td class="cost-value">${{ "%.2f"|format(service.cost) }}</td>
                    <td>{{ "%.1f"|format(service.percentage) }}%</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>

    <div class="section">
        <h2>Top Waste Categories</h2>
        <table>
            <thead>
                <tr><th>Category</th><th>Items</th><th>Potential Savings ($)</th></tr>
            </thead>
            <tbody>
                {% for category in waste_categories %}
                <tr>
                    <td>{{ category.name }}</td>
                    <td>{{ category.item_count }}</td>
                    <td class="savings">${{ "%.2f"|format(category.potential_savings) }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>

    <div class="section">
        <h2>Priority Recommendations</h2>
        {% for rec in recommendations %}
        <div class="metric">
            <h4>{{ rec.title }}</h4>
            <p>{{ rec.description }}</p>
            <p><strong>Potential Savings:</strong> <span class="savings">${{ "%.2f"|format(rec.estimated_savings) }}</span></p>
            <p><strong>Confidence:</strong> {{ rec.confidence }}%</p>
        </div>
        {% endfor %}
    </div>
</body>
</html>
            '''
        }

        self.jinja_env = Environment(loader=TemplateLoader(templates))

    async def generate_comprehensive_report(
        self,
        account_id: str,
        start_date: datetime,
        end_date: datetime,
        format_type: str = "pdf",
        user_id: str = None
    ) -> Dict[str, Any]:
        """Generate comprehensive cost optimization report"""

        logger.info("Generating comprehensive report",
                   account_id=account_id,
                   format_type=format_type)

        try:
            # Collect all data needed for the report
            report_data = await self._collect_report_data(account_id, start_date, end_date)

            # Generate report based on format
            if format_type.lower() == "pdf":
                file_path = await self._generate_pdf_report(account_id, report_data)
            elif format_type.lower() == "excel":
                file_path = await self._generate_excel_report(account_id, report_data)
            elif format_type.lower() == "html":
                file_path = await self._generate_html_report(account_id, report_data)
            else:
                raise ValueError(f"Unsupported format: {format_type}")

            # Cache report for quick access
            report_info = {
                "account_id": account_id,
                "format": format_type,
                "file_path": str(file_path),
                "generated_at": datetime.utcnow().isoformat(),
                "data_summary": {
                    "total_cost": report_data["summary"]["total_cost"],
                    "potential_savings": report_data["summary"]["potential_savings"],
                    "waste_items": len(report_data["waste_items"]),
                    "recommendations": len(report_data["recommendations"])
                }
            }

            cache_service.cache_report_info(account_id, report_info)

            return {
                "report_id": str(uuid.uuid4()),
                "file_path": str(file_path),
                "format": format_type,
                "size_bytes": file_path.stat().st_size,
                "generated_at": report_info["generated_at"],
                "summary": report_data["summary"]
            }

        except Exception as e:
            logger.error("Report generation failed", error=str(e), account_id=account_id)
            raise

    async def _collect_report_data(
        self,
        account_id: str,
        start_date: datetime,
        end_date: datetime
    ) -> Dict[str, Any]:
        """Collect all data needed for report generation"""

        async with get_database() as db:
            # Get account info
            from sqlalchemy import select
            result = await db.execute(
                select(AWSAccount).where(AWSAccount.id == account_id)
            )
            account = result.scalar_one_or_none()

            if not account:
                raise ValueError(f"Account not found: {account_id}")

            # Collect cost data
            cost_data = await cost_sync_service.get_cost_summary(
                account, start_date, end_date, db
            )

            # Collect waste data
            waste_data = await waste_detection_service.get_waste_summary(
                account, db
            )

            # Collect recommendations
            recommendations = await recommendations_engine.get_active_recommendations(
                account, db
            )

            # Calculate summary metrics
            total_cost = sum(item.get("cost", 0) for item in cost_data.get("daily_costs", []))
            potential_savings = sum(item.get("estimated_savings", 0) for item in waste_data.get("items", []))
            potential_savings += sum(rec.get("estimated_savings", 0) for rec in recommendations)

            # Process cost breakdown by service
            service_costs = {}
            for item in cost_data.get("service_costs", []):
                service = item.get("service_name", "Unknown")
                cost = item.get("cost", 0)
                service_costs[service] = service_costs.get(service, 0) + cost

            cost_breakdown = []
            for service, cost in sorted(service_costs.items(), key=lambda x: x[1], reverse=True):
                percentage = (cost / total_cost * 100) if total_cost > 0 else 0
                cost_breakdown.append({
                    "service_name": service,
                    "cost": cost,
                    "percentage": percentage
                })

            # Process waste categories
            waste_categories = {}
            for item in waste_data.get("items", []):
                category = item.get("category", "Other")
                if category not in waste_categories:
                    waste_categories[category] = {
                        "name": category,
                        "item_count": 0,
                        "potential_savings": 0
                    }
                waste_categories[category]["item_count"] += 1
                waste_categories[category]["potential_savings"] += item.get("estimated_savings", 0)

            return {
                "account": {
                    "id": str(account.id),
                    "name": account.name,
                    "account_id": account.aws_account_id
                },
                "period": {
                    "start_date": start_date.strftime("%Y-%m-%d"),
                    "end_date": end_date.strftime("%Y-%m-%d"),
                    "generated_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
                },
                "summary": {
                    "total_cost": total_cost,
                    "potential_savings": potential_savings,
                    "waste_items_count": len(waste_data.get("items", [])),
                    "recommendations_count": len(recommendations)
                },
                "cost_breakdown": cost_breakdown[:10],  # Top 10 services
                "waste_categories": list(waste_categories.values()),
                "waste_items": waste_data.get("items", [])[:20],  # Top 20 items
                "recommendations": recommendations[:10]  # Top 10 recommendations
            }

    async def _generate_pdf_report(self, account_id: str, data: Dict[str, Any]) -> Path:
        """Generate PDF report using ReportLab"""

        file_path = self.temp_dir / f"cost_report_{account_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

        # Create PDF document
        doc = SimpleDocTemplate(str(file_path), pagesize=letter)
        styles = getSampleStyleSheet()
        story = []

        # Title and header
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.darkblue,
            alignment=TA_CENTER,
            spaceAfter=30
        )

        story.append(Paragraph("AWS Cost Optimization Report", title_style))
        story.append(Paragraph(f"Account: {data['account']['name']} ({data['account']['account_id']})", styles['Heading2']))
        story.append(Paragraph(f"Period: {data['period']['start_date']} to {data['period']['end_date']}", styles['Normal']))
        story.append(Paragraph(f"Generated: {data['period']['generated_at']}", styles['Normal']))
        story.append(Spacer(1, 20))

        # Executive Summary
        story.append(Paragraph("Executive Summary", styles['Heading2']))

        summary_data = [
            ['Metric', 'Value'],
            ['Total Cost', f"${data['summary']['total_cost']:.2f}"],
            ['Potential Savings', f"${data['summary']['potential_savings']:.2f}"],
            ['Waste Items Found', str(data['summary']['waste_items_count'])],
            ['Active Recommendations', str(data['summary']['recommendations_count'])]
        ]

        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(summary_table)
        story.append(Spacer(1, 20))

        # Cost Breakdown by Service
        story.append(Paragraph("Cost Breakdown by Service", styles['Heading2']))

        cost_data = [['Service', 'Cost ($)', 'Percentage']]
        for service in data['cost_breakdown']:
            cost_data.append([
                service['service_name'],
                f"${service['cost']:.2f}",
                f"{service['percentage']:.1f}%"
            ])

        cost_table = Table(cost_data, colWidths=[3*inch, 1.5*inch, 1*inch])
        cost_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(cost_table)
        story.append(PageBreak())

        # Waste Categories
        story.append(Paragraph("Waste Detection Results", styles['Heading2']))

        waste_data = [['Category', 'Items', 'Potential Savings ($)']]
        for category in data['waste_categories']:
            waste_data.append([
                category['name'],
                str(category['item_count']),
                f"${category['potential_savings']:.2f}"
            ])

        waste_table = Table(waste_data, colWidths=[3*inch, 1*inch, 2*inch])
        waste_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(waste_table)
        story.append(Spacer(1, 20))

        # Top Recommendations
        story.append(Paragraph("Priority Recommendations", styles['Heading2']))

        for i, rec in enumerate(data['recommendations'][:5], 1):
            story.append(Paragraph(f"{i}. {rec.get('title', 'Recommendation')}", styles['Heading3']))
            story.append(Paragraph(rec.get('description', 'No description available'), styles['Normal']))
            story.append(Paragraph(f"Estimated Savings: ${rec.get('estimated_savings', 0):.2f}", styles['Normal']))
            story.append(Paragraph(f"Confidence: {rec.get('confidence_score', 0):.0f}%", styles['Normal']))
            story.append(Spacer(1, 10))

        # Build PDF
        doc.build(story)

        logger.info("PDF report generated successfully", file_path=str(file_path))
        return file_path

    async def _generate_excel_report(self, account_id: str, data: Dict[str, Any]) -> Path:
        """Generate Excel report using openpyxl"""

        file_path = self.temp_dir / f"cost_report_{account_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Create workbook
        wb = Workbook()

        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"

        # Header styling
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Title
        ws_summary['A1'] = "AWS Cost Optimization Report"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary['A2'] = f"Account: {data['account']['name']} ({data['account']['account_id']})"
        ws_summary['A3'] = f"Period: {data['period']['start_date']} to {data['period']['end_date']}"
        ws_summary['A4'] = f"Generated: {data['period']['generated_at']}"

        # Summary metrics
        ws_summary['A6'] = "Metric"
        ws_summary['B6'] = "Value"
        ws_summary['A6'].font = header_font
        ws_summary['A6'].fill = header_fill
        ws_summary['B6'].font = header_font
        ws_summary['B6'].fill = header_fill

        metrics = [
            ("Total Cost", f"${data['summary']['total_cost']:.2f}"),
            ("Potential Savings", f"${data['summary']['potential_savings']:.2f}"),
            ("Waste Items Found", data['summary']['waste_items_count']),
            ("Active Recommendations", data['summary']['recommendations_count'])
        ]

        for i, (metric, value) in enumerate(metrics, 7):
            ws_summary[f'A{i}'] = metric
            ws_summary[f'B{i}'] = value

        # Cost Breakdown Sheet
        ws_costs = wb.create_sheet("Cost Breakdown")
        ws_costs['A1'] = "Service"
        ws_costs['B1'] = "Cost ($)"
        ws_costs['C1'] = "Percentage"

        for i, col in enumerate(['A1', 'B1', 'C1']):
            ws_costs[col].font = header_font
            ws_costs[col].fill = header_fill

        for i, service in enumerate(data['cost_breakdown'], 2):
            ws_costs[f'A{i}'] = service['service_name']
            ws_costs[f'B{i}'] = service['cost']
            ws_costs[f'C{i}'] = f"{service['percentage']:.1f}%"

        # Add chart for cost breakdown
        if data['cost_breakdown']:
            chart = PieChart()
            labels = Reference(ws_costs, min_col=1, min_row=2, max_row=len(data['cost_breakdown'])+1)
            values = Reference(ws_costs, min_col=2, min_row=2, max_row=len(data['cost_breakdown'])+1)
            chart.add_data(values, titles_from_data=False)
            chart.set_categories(labels)
            chart.title = "Cost Distribution by Service"
            ws_costs.add_chart(chart, "E2")

        # Waste Categories Sheet
        ws_waste = wb.create_sheet("Waste Analysis")
        ws_waste['A1'] = "Category"
        ws_waste['B1'] = "Items"
        ws_waste['C1'] = "Potential Savings ($)"

        for i, col in enumerate(['A1', 'B1', 'C1']):
            ws_waste[col].font = header_font
            ws_waste[col].fill = header_fill

        for i, category in enumerate(data['waste_categories'], 2):
            ws_waste[f'A{i}'] = category['name']
            ws_waste[f'B{i}'] = category['item_count']
            ws_waste[f'C{i}'] = category['potential_savings']

        # Recommendations Sheet
        ws_recs = wb.create_sheet("Recommendations")
        ws_recs['A1'] = "Title"
        ws_recs['B1'] = "Description"
        ws_recs['C1'] = "Estimated Savings ($)"
        ws_recs['D1'] = "Confidence (%)"

        for i, col in enumerate(['A1', 'B1', 'C1', 'D1']):
            ws_recs[col].font = header_font
            ws_recs[col].fill = header_fill

        for i, rec in enumerate(data['recommendations'], 2):
            ws_recs[f'A{i}'] = rec.get('title', 'N/A')
            ws_recs[f'B{i}'] = rec.get('description', 'N/A')
            ws_recs[f'C{i}'] = rec.get('estimated_savings', 0)
            ws_recs[f'D{i}'] = rec.get('confidence_score', 0)

        # Auto-adjust column widths
        for ws in [ws_summary, ws_costs, ws_waste, ws_recs]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(str(file_path))

        logger.info("Excel report generated successfully", file_path=str(file_path))
        return file_path

    async def _generate_html_report(self, account_id: str, data: Dict[str, Any]) -> Path:
        """Generate HTML report using Jinja2 and WeasyPrint"""

        file_path = self.temp_dir / f"cost_report_{account_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"

        # Render HTML template
        template = self.jinja_env.get_template('cost_report.html')
        html_content = template.render(
            account_name=data['account']['name'],
            account_id=data['account']['account_id'],
            start_date=data['period']['start_date'],
            end_date=data['period']['end_date'],
            generated_at=data['period']['generated_at'],
            summary=data['summary'],
            cost_breakdown=data['cost_breakdown'],
            waste_categories=data['waste_categories'],
            recommendations=data['recommendations']
        )

        # Write HTML file
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        logger.info("HTML report generated successfully", file_path=str(file_path))
        return file_path

    async def schedule_periodic_reports(
        self,
        account_id: str,
        schedule_type: str = "monthly",
        format_types: List[str] = None,
        user_id: str = None
    ) -> Dict[str, Any]:
        """Schedule periodic report generation"""

        if format_types is None:
            format_types = ["pdf"]

        from app.services.queue_service import queue_service

        # Calculate next run time based on schedule
        now = datetime.utcnow()
        if schedule_type == "daily":
            next_run = now + timedelta(days=1)
            end_date = now.date()
            start_date = end_date - timedelta(days=7)  # Last 7 days
        elif schedule_type == "weekly":
            next_run = now + timedelta(weeks=1)
            end_date = now.date()
            start_date = end_date - timedelta(days=30)  # Last 30 days
        elif schedule_type == "monthly":
            next_run = now + timedelta(days=30)
            end_date = now.date()
            start_date = end_date - timedelta(days=90)  # Last 90 days
        else:
            raise ValueError(f"Invalid schedule type: {schedule_type}")

        # Schedule report generation job
        job_id = queue_service.enqueue_job(
            job_type="generate_scheduled_report",
            payload={
                "account_id": account_id,
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "format_types": format_types,
                "schedule_type": schedule_type,
                "user_id": user_id
            },
            delay=int((next_run - now).total_seconds())
        )

        return {
            "job_id": job_id,
            "schedule_type": schedule_type,
            "next_run": next_run.isoformat(),
            "format_types": format_types
        }

    async def get_report_status(self, report_id: str) -> Dict[str, Any]:
        """Get status of report generation"""

        # In a real implementation, you'd track report generation status
        # For now, return a mock status
        return {
            "report_id": report_id,
            "status": "completed",
            "progress": 100,
            "message": "Report generated successfully"
        }

    async def cleanup_old_reports(self, days_to_keep: int = 30) -> int:
        """Clean up old report files"""

        cutoff_date = datetime.now() - timedelta(days=days_to_keep)
        cleaned_count = 0

        try:
            for file_path in self.temp_dir.glob("cost_report_*.pdf"):
                if file_path.stat().st_mtime < cutoff_date.timestamp():
                    file_path.unlink()
                    cleaned_count += 1

            for file_path in self.temp_dir.glob("cost_report_*.xlsx"):
                if file_path.stat().st_mtime < cutoff_date.timestamp():
                    file_path.unlink()
                    cleaned_count += 1

            for file_path in self.temp_dir.glob("cost_report_*.html"):
                if file_path.stat().st_mtime < cutoff_date.timestamp():
                    file_path.unlink()
                    cleaned_count += 1

            logger.info("Old reports cleaned up", count=cleaned_count)

        except Exception as e:
            logger.error("Failed to clean up old reports", error=str(e))

        return cleaned_count


# Global report service instance
report_service = ReportService()